#!/usr/bin/env python
# coding=utf-8
from collections import namedtuple, OrderedDict
import json
import collections
from mobileStrings import csv_unicode
import os

__author__ = 'nic'

Wording = namedtuple("Wording", [
    'key',
    'exportable',
    'is_comment',
    'comment',
    'metadata',
    'translations'])
empty_wording = Wording('', True, '', False, OrderedDict(), OrderedDict())
create_wording = lambda **kwargs: empty_wording._replace(**kwargs)

FormatSpec = namedtuple('FormatSpec', [
    'key_col',
    'exportable_col',
    'is_comment_col',
    'comment_col',
    'translations_start_col',
    'exportable_rule',
    'is_comment_rule',
    'metadata_cols'])
default_format_specs = FormatSpec(0, 1, 2, 3, 4, bool, bool, {})
create_format_specs = lambda **kwargs: default_format_specs._replace(**kwargs)

def _get_csv_rows(file_path):
    csv_file = open(file_path, 'rb')
    reader = csv_unicode.UnicodeReader(csv_file)
    return reader

def _get_excel_rows(file_path, translations_start_col, sheet=0):
    return _get_excel_openpyxl_rows(file_path, sheet, translations_start_col)

def _get_excel_openpyxl_rows(file_path, sheet=0, translations_start_col=-1):
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=False, use_iterators=False)

    if hasattr(sheet, 'startswith'):
        work_sheet = wb.get_sheet_by_name(sheet)
    else:
        work_sheet = wb.worksheets[sheet] if sheet else wb.worksheets[0]

    iter_rows = work_sheet.iter_rows()

    header_row = iter_rows.next()

    if translations_start_col >= 0:
        header_row = list(header_row[:translations_start_col]) + \
                     list([v for v in header_row[translations_start_col:] if v.value])

    yield [v.value or '' for v in header_row]

    for row in iter_rows:
        yield [v.value or '' for v in row]

def _get_excel_xlrd_rows(file_path, sheet=0, translations_start_col=-1):
    import xlrd

    book = xlrd.open_workbook(file_path)
    sheet = book.sheets()[sheet]
    for row_index in range(sheet.nrows):
        row = sheet.row_values(row_index)
        yield row


def _check_duplicates(wordings, condition=lambda w: w.exportable and not w.is_comment,
                      message='ERROR: Duplicate key entry: "'):
    duplicate_keys = collections.defaultdict(lambda: [])

    for index, wording in enumerate(wordings):
        if condition(wording):
            duplicate_keys[wording.key].append(index)

    duplicate_keys = dict((k, v) for k, v in duplicate_keys.items() if len(v) > 1)

    for key, value in duplicate_keys.items():
        print(message + key + '" found at indices ' + ', '.join(str(v) for v in value))

    return duplicate_keys

def find_duplicate_wordings(wordings):
    return _check_duplicates(wordings)

def find_duplicate_comment_keys(wordings):
    return _check_duplicates(wordings, lambda w: w.is_comment, 'WARN: Duplicate comment key: "')

def fix_duplicates(wordings, merge_sections=True):
    new_wordings = wordings

    if find_duplicate_wordings(new_wordings):
        new_wordings = unique_wordings_overwrite(new_wordings)

    if merge_sections and find_duplicate_comment_keys(new_wordings):
        new_wordings = group_wordings_by_comment_key(new_wordings)

    return new_wordings

def trim(wordings):
    for w in wordings:
        for lang, t in w.translations.items():
            if hasattr(t, 'strip'):
                w.translations[lang] = t.strip()

def _read_rows(reader, specs=default_format_specs):
    languages = reader.next()[specs.translations_start_col:]
    wordings = list()

    for row_values in reader:
        if row_values:
            w = create_wording(
                key=row_values[specs.key_col].strip(),
                exportable=specs.exportable_rule(row_values[specs.exportable_col]),
                comment=row_values[specs.comment_col],
                is_comment=specs.is_comment_rule(row_values[specs.is_comment_col]),
                metadata=OrderedDict(((k, row_values[v]) for k, v in specs.metadata_cols.items())),
                translations=OrderedDict(zip(languages, [v for v in row_values[specs.translations_start_col:]]))
            )

        wordings.append(w)

    return languages, wordings

def group_wordings_by_comment_key(wordings):
    grouped_wordings = OrderedDict()  # comment_key, [wording, ...]
    comment_keys = OrderedDict()  # comment_key, comment_wording
    current_comment_key = None

    for wording in wordings:
        if wording.is_comment:
            current_comment_key = wording.key
            grouped_wordings.setdefault(current_comment_key, [])
            comment_keys[wording.key] = wording
        else:
            grouped_wordings.setdefault(current_comment_key, []).append(wording)

    new_wordings = []
    for k, v in grouped_wordings.items():
        if k in comment_keys:
            new_wordings.append(comment_keys.get(k))
        for w in v:
            new_wordings.append(w)

    return new_wordings

def unique_wordings_overwrite(wordings):
    new_wordings = OrderedDict()
    duplicate_other = collections.defaultdict(lambda: [])

    for index, wording in enumerate(wordings):
        if wording.exportable and not wording.is_comment:
            new_wordings[wording.key] = wording
        else:
            # Ignore duplicate keys if not exportable
            new_key = wording.key+'.'+str(len(duplicate_other[wording.key]))
            new_wordings[new_key] = wording
            duplicate_other[wording.key].append(index+2)

    return new_wordings.values()


def _object_hook(dct):
    if dct and dct[0] and dct[0][0] == 'key':
        keys = [k for k, v in dct]
        if keys[-1] == 'translations':
            return create_wording(**dict((k, v) for k, v in dct if k in Wording._fields))
    return OrderedDict(dct)

def _read_json(file__obj):
    dct = json.load(file__obj, 'utf-8', object_pairs_hook=_object_hook)
    return dct['languages'], dct['wordings']

def read_json(file_or_path):
    if hasattr(file_or_path, 'write'):
        return _read_json(file_or_path)
    else:
        with open(file_or_path, 'r') as f:
            return _read_json(f)

def read_excel(file_path, rows_format_specs=default_format_specs, sheet=0):
    return _read_rows(_get_excel_rows(file_path, rows_format_specs.translations_start_col, sheet), rows_format_specs)


def read_csv(file_path, rows_format_specs=default_format_specs):
    return _read_rows(_get_csv_rows(file_path), rows_format_specs)


def read_file(file_path, rows_format_specs=default_format_specs):

    _, ext = os.path.splitext(file_path.lower())

    if ext == '.json':
        return read_json(file_path)
    elif ext == '.csv':
        return read_csv(file_path, rows_format_specs)
    elif ext.startswith('.xls'):
        return read_excel(file_path, rows_format_specs)
    else:
        raise AttributeError("Unknown file type: " + ext)
