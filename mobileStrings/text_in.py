#!/usr/bin/env python
# coding=utf-8

from collections import namedtuple, OrderedDict
import simplejson as json
import collections
from mobileStrings import csv_unicode
from mobileStrings.collection_utils import namedtuple_with_defaults
import os

__author__ = 'nic'

Wordings = namedtuple("Wordings", """\
    languages,
    wordings""")

Wording = namedtuple_with_defaults("Wording", """\
    key,
    exportable,
    is_comment,
    comment,
    metadata,
    translations""", default_values=dict(key='', exportable=True, is_comment=False, comment='',
                                         metadata=OrderedDict(), translations=OrderedDict()))

FormatSpec = namedtuple_with_defaults('FormatSpec', """\
    excel_sheet_reference,
    key_col,
    exportable_col,
    is_comment_col,
    comment_col,
    translations_start_col,
    exportable_value,
    is_comment_value,
    metadata_cols""", default_values=dict(excel_sheet_reference=0,
                                          key_col=0, exportable_col=1, is_comment_col=2,
                                          comment_col=3, translations_start_col=4,
                                          exportable_value=None, is_comment_value=None,
                                          metadata_cols={}))

default_format_specs = FormatSpec()


def _get_csv_rows(file_path):
    csv_file = open(file_path, 'rb')
    reader = csv_unicode.UnicodeReader(csv_file)
    return reader


def _get_excel_rows(file_path, translations_start_col, sheet=0):
    return _get_excel_openpyxl_rows(file_path, sheet, translations_start_col)


def _get_excel_openpyxl_rows(file_path, sheet=0, translations_start_col=-1):
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=False, use_iterators=False)

    if hasattr(sheet, 'startswith'):
        work_sheet = wb.get_sheet_by_name(sheet)
    else:
        work_sheet = wb.worksheets[sheet] if sheet else wb.worksheets[0]

    iter_rows = work_sheet.iter_rows()

    header_row = iter_rows.next()

    if translations_start_col >= 0:
        header_row = list(header_row[:translations_start_col]) + \
                     list([v for v in header_row[translations_start_col:] if v.value])

    yield [v.value or '' for v in header_row]

    for row in iter_rows:
        yield [v.value and unicode(v.value).replace(u'\u2028', '\n') or '' for v in row]


def _check_duplicates(wordings, condition=lambda w: w.exportable and not w.is_comment,
                      message='WARN: Duplicate key entry overwritten: "'):
    duplicate_keys = collections.defaultdict(lambda: [])

    for index, wording in enumerate(wordings):
        if condition(wording):
            duplicate_keys[wording.key].append(index)

    duplicate_keys = dict((k, v) for k, v in duplicate_keys.items() if len(v) > 1)

    for key, value in duplicate_keys.items():
        print(message + key + '" found at indices ' + ', '.join(str(v) for v in value))

    return duplicate_keys


def find_duplicate_wordings(wordings):
    return _check_duplicates(wordings)


def find_duplicate_comment_keys(wordings):
    return _check_duplicates(wordings, lambda w: w.is_comment,
                             'INFO: Merged duplicate section for comment key: "')


def fix_duplicates(wordings, merge_sections=True):
    new_wordings = wordings

    if find_duplicate_wordings(new_wordings):
        new_wordings = unique_wordings_overwrite(new_wordings)

    if merge_sections and find_duplicate_comment_keys(new_wordings):
        new_wordings = group_wordings_by_comment_key(new_wordings)

    return new_wordings


def _read_rows(reader, specs=default_format_specs):
    languages = reader.next()[specs.translations_start_col:]
    wordings = _wordings_generator(languages, reader, specs)
    return languages, wordings


def _bool_in(something):
    if something is None:
        return bool
    if isinstance(something, (list, tuple)):
        return lambda v: v in something
    else:
        return lambda v: v == something


def _wordings_generator(languages, reader, specs):

    exportable_rule = _bool_in(specs.exportable_value)
    is_comment_rule = _bool_in(specs.is_comment_value)

    for row_values in reader:
        if row_values:
            w = Wording(
                key=row_values[specs.key_col].strip(),
                exportable=exportable_rule(row_values[specs.exportable_col]),
                comment=row_values[specs.comment_col],
                is_comment=is_comment_rule(row_values[specs.is_comment_col]),
                metadata=OrderedDict(((k, row_values[v]) for k, v in specs.metadata_cols.items())),
                translations=OrderedDict(
                    zip(languages, [v for v in row_values[specs.translations_start_col:]]))
            )

            yield w


def group_wordings_by_comment_key(wordings):
    grouped_wordings = OrderedDict()  # comment_key, [wording, ...]
    comment_keys = OrderedDict()  # comment_key, comment_wording
    current_comment_key = None

    for wording in wordings:
        if wording.is_comment:
            current_comment_key = wording.key
            grouped_wordings.setdefault(current_comment_key, [])
            comment_keys[wording.key] = wording
        else:
            grouped_wordings.setdefault(current_comment_key, []).append(wording)

    new_wordings = []
    for k, v in grouped_wordings.items():
        if k in comment_keys:
            new_wordings.append(comment_keys.get(k))
        for w in v:
            new_wordings.append(w)

    return new_wordings


def unique_wordings_overwrite(wordings):
    new_wordings = OrderedDict()
    duplicate_other = collections.defaultdict(lambda: [])

    for index, wording in enumerate(wordings):
        if wording.exportable and not wording.is_comment:
            new_wordings[wording.key] = wording
        else:
            # Ignore duplicate keys if not exportable
            new_key = wording.key+'.'+str(len(duplicate_other[wording.key]))
            new_wordings[new_key] = wording
            duplicate_other[wording.key].append(index+2)

    return new_wordings.values()


def are_keys_from_list(dct, fields):
    return len([k for k, v in dct if k in fields]) == len(dct)


def _object_hook(dct):
    if dct:
        if are_keys_from_list(dct, Wording._fields):
            return Wording(**dict(dct))

        if are_keys_from_list(dct, Wordings._fields):
            return Wordings(**dict(dct))

    return OrderedDict(dct)


def _read_json(file__obj):
    return json.load(file__obj, 'utf-8', object_pairs_hook=_object_hook)


def read_json(file_or_path):
    if hasattr(file_or_path, 'read'):
        return _read_json(file_or_path)
    else:
        with open(file_or_path, 'r') as f:
            return _read_json(f)


def iread_excel(file_path, rows_format_specs=default_format_specs):
    return _read_rows(
            _get_excel_rows(file_path,
                            rows_format_specs.translations_start_col,
                            rows_format_specs.excel_sheet_reference),
            rows_format_specs)


def read_excel(file_path, rows_format_specs=default_format_specs):
    languages, wordings = iread_excel(file_path, rows_format_specs)
    return languages, list(wordings)


def iread_csv(file_path, rows_format_specs=default_format_specs):
    return _read_rows(_get_csv_rows(file_path), rows_format_specs)


def read_csv(file_path, rows_format_specs=default_format_specs):
    languages, wordings = _read_rows(_get_csv_rows(file_path), rows_format_specs)
    return languages, list(wordings)


def _config_object_pairs_hook(dct):
    if dct and are_keys_from_list(dct, FormatSpec._fields):
        return FormatSpec(**dict(dct))
    return OrderedDict(dct)


def read_row_format_config(config_file):
    rows_format_specs = default_format_specs
    if config_file:
        with open(config_file, 'r') as f:
            rows_format_specs = json.load(f, object_pairs_hook=_config_object_pairs_hook)
    return rows_format_specs


def read_file(file_path, rows_format_specs=default_format_specs, prefer_generator=True):

    _, ext = os.path.splitext(file_path.lower())

    if ext == '.json':
        return read_json(file_path)
    elif ext == '.csv':
        read = iread_csv if prefer_generator else read_csv
        return read(file_path, rows_format_specs)
    elif ext.startswith('.xls'):
        read = iread_excel if prefer_generator else read_excel
        return read(file_path, rows_format_specs)
    else:
        raise AttributeError("Unknown file type: " + ext)
